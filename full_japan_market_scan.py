# full_japan_market_scan.py
# =========================
# Tokyo Stock Exchange (TSE) universe scanner using kabupy + yfinance.
#
# Pipeline:
#   Stage 1 — Fetch TSE Prime / Standard equity universe via kabupy (+ JPX CSV fallback)
#   Stage 2 — Bulk OHLC download via yfinance (tickers suffixed with .T)
#   Stage 3 — Darvas Box screen + 200-day MA trend on every stock
#   Stage 4 — Piotroski F-Score + Coffee Can on Darvas BREAKOUT candidates only
#   Stage 5 — Save styled Excel workbook with ranked results
#
# Output sheets:
#   All_Stocks      — price summary for every stock scanned
#   Darvas_Signals  — breakout / breakdown alerts ranked by upside
#   Fundamentals    — Piotroski + Coffee Can results for breakout candidates
#   Triple_Hits     — BREAKOUT_BUY + Piotroski ≥ 7 + Coffee Can PASS
#
# Usage:
#   python full_japan_market_scan.py
#   python full_japan_market_scan.py --top 200     # limit to first 200 tickers
#   python full_japan_market_scan.py --no-scans    # Darvas only, skip fundamentals
#   python full_japan_market_scan.py --workers 8
#
# Install:
#   pip install kabupy yfinance pandas openpyxl requests

import argparse
import json
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

import pandas as pd
import requests
import yfinance as yf

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Configuration ─────────────────────────────────────────────────────────────
DOWNLOAD_DIR     = Path("./japan_scan")
DOWNLOAD_DIR.mkdir(exist_ok=True)

DARVAS_CONFIRM      = 3
BATCH_SIZE          = 200
SLEEP_BETWEEN       = 1.5
MAX_WORKERS         = 10
PIOTROSKI_STRONG    = 7
MAX_FUND_CANDIDATES = 200
SYMBOL_CACHE_TTL    = 86400   # 24 h

_CACHE_FILE = DOWNLOAD_DIR / ".symbols_cache.json"

# JPX publishes a daily stock master CSV (TSE Prime + Standard + Growth).
# This is the authoritative free source for the full TSE universe.
JPX_MASTER_URL = (
    "https://www.jpx.co.jp/markets/statistics-equities/misc/"
    "tvdivq0000001vg2-att/data_j.xls"
)

# ── Symbol universe ───────────────────────────────────────────────────────────

def _load_cache():
    try:
        if _CACHE_FILE.exists():
            data = json.loads(_CACHE_FILE.read_text())
            if time.time() - data.get("ts", 0) < SYMBOL_CACHE_TTL:
                return data.get("tickers", [])
    except Exception:
        pass
    return None


def _save_cache(tickers):
    try:
        _CACHE_FILE.write_text(json.dumps({"ts": time.time(), "tickers": tickers}))
    except Exception:
        pass


def fetch_tse_universe_kabupy() -> list[dict]:
    """Fetch TSE equity universe via kabupy.Jpx().issues (official JPX data)."""
    try:
        import kabupy
        issues = kabupy.Jpx().issues   # list of dicts, ~4000+ records
        result = []
        for s in issues:
            cat = str(s.get("category", ""))
            # Keep Prime and Standard domestic equities; skip ETF/REIT/foreign
            if "内国株式" not in cat:
                continue
            code = str(s.get("security_code", "")).split(".")[0].zfill(4)
            if not code.isdigit() or len(code) != 4:
                continue
            result.append({
                "code":   code,
                "name":   s.get("name", ""),
                "sector": s.get("33_industry_category", s.get("17_industry_category", "—")),
                "market": cat.replace("（内国株式）", ""),
            })
        if result:
            print(f"  kabupy → {len(result)} TSE domestic equities")
        return result
    except ImportError:
        print("  kabupy not installed — falling back to JPX master CSV")
    except Exception as e:
        print(f"  kabupy error: {e} — falling back to JPX master CSV")
    return []


def fetch_tse_universe_jpx() -> list[dict]:
    """Download TSE stock master Excel from JPX (no auth required, free)."""
    print("  Fetching TSE universe from JPX stock master …")
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}
        r = requests.get(JPX_MASTER_URL, headers=headers, timeout=30)
        r.raise_for_status()
        df = pd.read_excel(BytesIO(r.content), dtype=str)
        # Expected columns vary by JPX file version; normalise
        df.columns = [c.strip() for c in df.columns]
        code_col   = next((c for c in df.columns if "コード" in c or "Code" in c), None)
        name_col   = next((c for c in df.columns if "銘柄名" in c or "銘柄" in c or "Name" in c.capitalize()), None)
        sector_col = next((c for c in df.columns if "業種" in c or "Sector" in c.capitalize()), None)
        market_col = next((c for c in df.columns if "市場" in c or "Market" in c.capitalize()), None)

        if code_col is None:
            print("  ⚠️  Cannot identify code column in JPX file")
            return []

        # Filter to Prime + Standard markets (equities only)
        if market_col:
            df = df[df[market_col].str.contains("Prime|Standard|プライム|スタンダード", na=False)]

        result = []
        for _, row in df.iterrows():
            code = str(row[code_col]).strip().split(".")[0].zfill(4)
            if not code.isdigit() or len(code) != 4:
                continue
            result.append({
                "code":   code,
                "name":   str(row[name_col]).strip() if name_col else "",
                "sector": str(row[sector_col]).strip() if sector_col else "—",
            })
        return result
    except Exception as e:
        print(f"  ⚠️  JPX fetch error: {e}")
        return []


def build_ticker_universe() -> list[dict]:
    """Return deduplicated TSE ticker list with name/sector, cached for 24 h."""
    cached = _load_cache()
    if cached:
        print(f"  Symbol cache hit: {len(cached)} TSE tickers")
        return cached

    stocks = fetch_tse_universe_kabupy()
    if not stocks:
        stocks = fetch_tse_universe_jpx()

    if not stocks:
        sys.exit("❌  Could not fetch TSE universe from kabupy or JPX. Check network/install.")

    # Deduplicate on code, build yfinance ticker (.T suffix)
    seen = set()
    result = []
    for s in stocks:
        code = s["code"]
        if code in seen:
            continue
        seen.add(code)
        result.append({
            "yf_ticker": f"{code}.T",
            "code":      code,
            "name":      s["name"],
            "sector":    s["sector"],
        })

    _save_cache(result)
    print(f"  → {len(result)} unique TSE equity tickers")
    return result


# ── Bulk OHLC ─────────────────────────────────────────────────────────────────

def bulk_download_ohlc(tickers: list[str], period: str = "3mo") -> dict[str, pd.DataFrame]:
    result: dict[str, pd.DataFrame] = {}
    batches = [tickers[i:i + BATCH_SIZE] for i in range(0, len(tickers), BATCH_SIZE)]
    print(f"  Downloading OHLC for {len(tickers)} tickers in {len(batches)} batches …")

    for idx, batch in enumerate(batches, 1):
        print(f"    Batch {idx}/{len(batches)} ({len(batch)} tickers) …", end=" ", flush=True)
        try:
            raw = yf.download(batch, period=period, auto_adjust=True,
                              threads=True, progress=False)
            if raw.empty:
                print("empty")
                continue
            if isinstance(raw.columns, pd.MultiIndex):
                for tkr in batch:
                    try:
                        df = raw.xs(tkr, axis=1, level=1).dropna(how="all")
                        if not df.empty and len(df) >= DARVAS_CONFIRM + 5:
                            result[tkr] = df
                    except KeyError:
                        pass
            else:
                tkr = batch[0]
                if not raw.empty:
                    result[tkr] = raw
            print(f"OK ({sum(1 for t in batch if t in result)} usable)")
        except Exception as e:
            print(f"ERROR — {e}")
        if idx < len(batches):
            time.sleep(SLEEP_BETWEEN)

    return result


# ── Darvas Box ────────────────────────────────────────────────────────────────

def compute_darvas_box(df: pd.DataFrame, confirm: int = DARVAS_CONFIRM) -> dict:
    if df is None or df.empty or len(df) < confirm + 5:
        return {"signal": "INSUFFICIENT_DATA", "box_top": None, "box_bottom": None}

    highs  = pd.to_numeric(df["High"],  errors="coerce").fillna(0).tolist()
    lows   = pd.to_numeric(df["Low"],   errors="coerce").fillna(0).tolist()
    closes = pd.to_numeric(df["Close"], errors="coerce").fillna(0).tolist()

    current = closes[-1]
    h = highs[:-1]   # exclude current bar
    l = lows[:-1]
    n = len(h)

    box_top_idx = box_top = None
    for i in range(n - confirm - 1, -1, -1):
        c = h[i]
        if c == 0:
            continue
        w = h[i + 1: i + 1 + confirm]
        if len(w) == confirm and all(x < c for x in w):
            box_top_idx, box_top = i, c
            break

    if box_top is None:
        return {"signal": "NO_BOX", "box_top": None, "box_bottom": None, "current_price": current}

    seg = l[box_top_idx:]
    box_bottom = None
    for i in range(len(seg) - confirm):
        c = seg[i]
        if c == 0:
            continue
        w = seg[i + 1: i + 1 + confirm]
        if len(w) == confirm and all(x > c for x in w):
            box_bottom = c
            break
    if box_bottom is None:
        valid = [x for x in seg if x > 0]
        box_bottom = min(valid) if valid else None

    if box_bottom is None:
        return {"signal": "NO_BOX", "box_top": round(box_top, 0), "box_bottom": None,
                "current_price": round(current, 0)}

    signal = ("BREAKOUT_BUY"   if current > box_top   else
              "BREAKDOWN_SELL" if current < box_bottom else "IN_BOX")

    rng    = box_top - box_bottom
    upside = ((box_top - current) / current * 100) if current else 0
    pos    = ((current - box_bottom) / rng * 100)  if rng    else 0

    return {
        "signal":       signal,
        "box_top":      round(box_top,    0),
        "box_bottom":   round(box_bottom, 0),
        "current_price":round(current,    0),
        "upside_pct":   round(upside, 2),
        "pos_in_box":   round(pos,    1),
        "data_points":  len(closes),
    }


# ── Fundamental helpers ───────────────────────────────────────────────────────

def _first_df(ticker, *attrs):
    for attr in attrs:
        df = getattr(ticker, attr, None)
        if df is not None and isinstance(df, pd.DataFrame) and not df.empty:
            return df
    return None


def _row(df, *names, col: int = 0):
    if df is None or df.empty:
        return None
    for name in names:
        if name in df.index:
            try:
                val = df.loc[name].iloc[col]
                return float(val) if pd.notna(val) else None
            except Exception:
                pass
    return None


def _series(df, *names):
    if df is None or df.empty:
        return []
    for name in names:
        if name in df.index:
            return [float(v) for v in df.loc[name].dropna() if pd.notna(v)]
    return []


# ── Fundamental scan ─────────────────────────────────────────────────────────

def fundamental_scan(yf_ticker: str) -> dict:
    """Piotroski F-Score + Japan-adapted Coffee Can in a single yfinance call."""
    result = {"yf_ticker": yf_ticker, "f_score": None, "cc_qualifies": "FAIL", "error": None}
    try:
        ticker = yf.Ticker(yf_ticker)
        inc = _first_df(ticker, "income_stmt", "financials")
        bal = _first_df(ticker, "balance_sheet")
        cf  = _first_df(ticker, "cash_flow", "cashflow")
        if inc is None:
            result["error"] = "no_financial_data"
            return result
    except Exception as e:
        result["error"] = str(e)[:80]
        return result

    # ── Piotroski F-Score ─────────────────────────────────────────────────────
    try:
        ni0 = _row(inc, "Net Income", col=0);  a0 = _row(bal, "Total Assets", col=0)
        ni1 = _row(inc, "Net Income", col=1);  a1 = _row(bal, "Total Assets", col=1)
        roa0 = (ni0 / a0) if (ni0 and a0) else None
        roa1 = (ni1 / a1) if (ni1 and a1) else None
        ocf0 = _row(cf, "Operating Cash Flow", "Total Cash From Operating Activities")

        f1 = 1 if (roa0 and roa0 > 0)              else 0
        f2 = 1 if (ocf0 and ocf0 > 0)              else 0
        f3 = 1 if (roa0 and roa1 and roa0 > roa1)  else 0
        f4 = 1 if (ocf0 and a0 and roa0 and (ocf0 / a0) > roa0) else 0

        ltd0 = _row(bal, "Long Term Debt", col=0) or 0
        ltd1 = _row(bal, "Long Term Debt", col=1) or 0
        f5 = 1 if (a0 and a1 and (ltd0 / a0) < (ltd1 / a1)) else 0

        ca0 = _row(bal, "Current Assets", "Total Current Assets", col=0)
        cl0 = _row(bal, "Current Liabilities", "Total Current Liabilities", col=0)
        ca1 = _row(bal, "Current Assets", "Total Current Assets", col=1)
        cl1 = _row(bal, "Current Liabilities", "Total Current Liabilities", col=1)
        f6 = 1 if (ca0 and cl0 and ca1 and cl1 and (ca0 / cl0) > (ca1 / cl1)) else 0

        sh0 = _row(bal, "Share Issued", col=0)
        sh1 = _row(bal, "Share Issued", col=1)
        f7 = (1 if sh0 <= sh1 else 0) if (sh0 and sh1) else 1

        rev0 = _row(inc, "Total Revenue", col=0); gp0 = _row(inc, "Gross Profit", col=0)
        rev1 = _row(inc, "Total Revenue", col=1); gp1 = _row(inc, "Gross Profit", col=1)
        f8 = 1 if (gp0 and rev0 and gp1 and rev1 and (gp0 / rev0) > (gp1 / rev1)) else 0
        f9 = 1 if (rev0 and a0 and rev1 and a1 and (rev0 / a0) > (rev1 / a1)) else 0

        f_score = f1 + f2 + f3 + f4 + f5 + f6 + f7 + f8 + f9
        result["f_score"]  = f_score
        result["f_strong"] = f_score >= PIOTROSKI_STRONG
    except Exception:
        return result

    if not result.get("f_strong"):
        result["cc_score"] = "—"
        return result

    # ── Coffee Can (Japan-adapted) ────────────────────────────────────────────
    # C1: Revenue CAGR > 10% over available history
    # C2: Avg ROCE > 15%
    # C3: D/E < 1
    # C4: Market cap ≥ ¥100B (~mid-cap TSE Prime)
    # C5: Consistently profitable (all years positive NI)
    # C6: Positive FCF (most recent year)
    try:
        c = {}
        revs = _series(inc, "Total Revenue")
        if len(revs) >= 2 and revs[-1] > 0:
            cagr = ((revs[0] / revs[-1]) ** (1 / (len(revs) - 1)) - 1) * 100
            c["C1"] = 1 if cagr > 10 else 0
        else:
            cagr = None; c["C1"] = 0

        ebit_s = _series(inc, "EBIT", "Operating Income", "Ebit")
        ta_s   = _series(bal, "Total Assets")
        cl_s   = _series(bal, "Current Liabilities", "Total Current Liabilities")
        roce_l = [ebit_s[i] / (ta_s[i] - cl_s[i]) * 100
                  for i in range(min(len(ebit_s), len(ta_s), len(cl_s)))
                  if (ta_s[i] - cl_s[i]) > 0]
        avg_roce = sum(roce_l) / len(roce_l) if roce_l else None
        c["C2"] = 1 if (avg_roce and avg_roce > 15) else 0

        ltd_s = _series(bal, "Long Term Debt")
        eq_s  = _series(bal, "Stockholders Equity", "Total Stockholder Equity",
                        "Total Equity Gross Minority Interest")
        de = (ltd_s[0] / abs(eq_s[0])) if (ltd_s and eq_s and eq_s[0] != 0) else None
        c["C3"] = 1 if (de is not None and de < 1) else 0

        try:
            mcap = ticker.fast_info.market_cap or 0
        except Exception:
            mcap = 0
        c["C4"] = 1 if mcap >= 1e11 else 0   # ≥ ¥100B

        ni_s = _series(inc, "Net Income")
        c["C5"] = 1 if (ni_s and all(n > 0 for n in ni_s)) else 0

        fcf_s = _series(cf, "Free Cash Flow")
        if fcf_s:
            c["C6"] = 1 if fcf_s[0] > 0 else 0
        else:
            ocf_s   = _series(cf, "Operating Cash Flow")
            capex_s = _series(cf, "Capital Expenditure")
            c["C6"] = 1 if (ocf_s and capex_s and (ocf_s[0] - abs(capex_s[0])) > 0) else 0

        qualifies = sum(c.values()) == len(c)
        result.update({
            "cc_qualifies":   "PASS" if qualifies else "FAIL",
            "cc_score":       f"{sum(c.values())}/{len(c)}",
            "cc_rev_cagr":    round(cagr, 2) if cagr else None,
            "cc_roce_avg":    round(avg_roce, 2) if avg_roce else None,
            "cc_debt_equity": round(de, 2) if de is not None else None,
            "market_cap_b_jpy": round(mcap / 1e9, 1) if mcap else None,
        })
    except Exception:
        pass

    return result


# ── Excel styling ─────────────────────────────────────────────────────────────

def style_sheet(ws):
    if not OPENPYXL_OK:
        return
    fill_hdr  = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    fill_alt  = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    font_hdr  = Font(name="Calibri", size=11, bold=True,  color="FFFFFF")
    font_body = Font(name="Calibri", size=11, bold=False, color="000000")
    thin = Border(
        left=Side(style="thin", color="CBD5E0"), right=Side(style="thin", color="CBD5E0"),
        top=Side(style="thin", color="CBD5E0"),  bottom=Side(style="thin", color="CBD5E0"),
    )
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_hdr; cell.fill = fill_hdr; cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_body; cell.border = thin
            if row_idx % 2 == 1:
                cell.fill = fill_alt
            hdr = str(ws.cell(row=1, column=col_idx).value or "").upper()
            val = cell.value
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if any(k in hdr for k in ["%", "CAGR", "ROCE", "YIELD"]):
                    cell.number_format = '0.00"%"'
                elif any(k in hdr for k in ["CAP", "LTP", "BOX", "PRICE", "200"]):
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '0.00'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Full TSE scanner — Darvas + Piotroski + Coffee Can")
    parser.add_argument("--top",      type=int, default=0,     help="Limit to first N tickers")
    parser.add_argument("--no-scans", action="store_true",     help="Skip fundamental scans")
    parser.add_argument("--workers",  type=int, default=MAX_WORKERS, help="Parallel threads")
    args = parser.parse_args()

    print(f"\n{'#'*60}")
    print(f"  FULL JAPAN (TSE) MARKET SCAN")
    print(f"  Started: {datetime.now().strftime('%d %b %Y  %H:%M:%S')}")
    print(f"{'#'*60}\n")

    # Stage 1 — Universe
    print("Stage 1 — Building TSE equity universe …")
    universe = build_ticker_universe()
    meta = {s["yf_ticker"]: s for s in universe}
    all_yf_tickers = [s["yf_ticker"] for s in universe]
    if args.top:
        all_yf_tickers = all_yf_tickers[:args.top]
        print(f"  (limited to first {args.top} tickers)")

    # Stage 2 — Bulk OHLC
    print(f"\nStage 2 — Bulk OHLC download ({len(all_yf_tickers)} tickers) …")
    ohlc = bulk_download_ohlc(all_yf_tickers, period="3mo")
    print(f"  → {len(ohlc)} tickers with usable OHLC data")

    # Stage 3 — Darvas Box + 200-day MA
    print("\nStage 3 — Darvas Box & 200-day MA screen …")
    all_rows, darvas_rows, breakout_tickers = [], [], []

    for yf_tkr, df in ohlc.items():
        info    = meta.get(yf_tkr, {})
        darv    = compute_darvas_box(df)
        closes  = pd.to_numeric(df["Close"], errors="coerce").dropna()
        ltp     = round(float(closes.iloc[-1]), 0)  if not closes.empty  else None
        prev    = round(float(closes.iloc[-2]), 0)  if len(closes) >= 2  else None
        chg_pct = round((ltp - prev) / prev * 100, 2) if (ltp and prev) else None

        ma200   = round(closes.rolling(200).mean().iloc[-1], 0) if len(closes) >= 200 else None
        dist_ma = round((ltp - ma200) / ma200 * 100, 2)         if (ma200 and ltp)   else None
        trend   = ("Above 200MA (Uptrend)"    if dist_ma and dist_ma >  5 else
                   "Below 200MA (Downtrend)"  if dist_ma and dist_ma < -5 else
                   "Near 200MA (Consolidation)" if dist_ma else "Insufficient History")

        row = {
            "YF_Ticker":          yf_tkr,
            "Code":               info.get("code", yf_tkr.replace(".T", "")),
            "Name":               info.get("name", ""),
            "Sector":             info.get("sector", "—"),
            "LTP_JPY":            ltp,
            "Change%":            chg_pct,
            "200_Day_MA":         ma200,
            "Distance_to_200MA%": dist_ma,
            "Trend_Signal":       trend,
            "Darvas_Signal":      darv.get("signal"),
            "Box_Top":            darv.get("box_top"),
            "Box_Bottom":         darv.get("box_bottom"),
            "Upside_to_Top%":     darv.get("upside_pct"),
            "Position_in_Box%":   darv.get("pos_in_box"),
            "Data_Points":        darv.get("data_points"),
        }
        all_rows.append(row)
        if darv.get("signal") in ("BREAKOUT_BUY", "BREAKDOWN_SELL"):
            darvas_rows.append(row.copy())
        if darv.get("signal") == "BREAKOUT_BUY":
            breakout_tickers.append(yf_tkr)

    print(f"  Breakout BUY:   {len(breakout_tickers)}")
    print(f"  Breakdown SELL: {sum(1 for r in darvas_rows if r['Darvas_Signal'] == 'BREAKDOWN_SELL')}")
    print(f"  In Box:         {len(all_rows) - len(darvas_rows)}")

    # Stage 4 — Fundamentals on breakout candidates
    fund_rows, triple_rows = [], []

    if not args.no_scans and breakout_tickers:
        # Cap to freshest MAX_FUND_CANDIDATES breakouts
        def _upside(tkr):
            row = next((r for r in darvas_rows if r["YF_Ticker"] == tkr), None)
            return abs(row.get("Upside_to_Top%") or 999)
        if len(breakout_tickers) > MAX_FUND_CANDIDATES:
            breakout_tickers = sorted(breakout_tickers, key=_upside)[:MAX_FUND_CANDIDATES]
            print(f"  (capped to {MAX_FUND_CANDIDATES} freshest breakouts)")

        print(f"\nStage 4 — Fundamental scans ({len(breakout_tickers)} candidates, "
              f"{args.workers} workers) …")
        done = 0
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            futures = {pool.submit(fundamental_scan, tkr): tkr for tkr in breakout_tickers}
            for future in as_completed(futures):
                tkr  = futures[future]
                done += 1
                try:
                    res  = future.result()
                    tech = next((r for r in darvas_rows if r["YF_Ticker"] == tkr), {})
                    fund_row = {
                        "YF_Ticker":       tkr,
                        "Code":            tech.get("Code"),
                        "Name":            tech.get("Name"),
                        "Sector":          tech.get("Sector"),
                        "LTP_JPY":         tech.get("LTP_JPY"),
                        "Change%":         tech.get("Change%"),
                        "Darvas_Signal":   tech.get("Darvas_Signal"),
                        "Upside_to_Top%":  tech.get("Upside_to_Top%"),
                        "200_Day_MA":      tech.get("200_Day_MA"),
                        "Piotroski_Score": res.get("f_score"),
                        "Piotroski_Strong":("YES" if res.get("f_strong") else "NO"),
                        "CoffeeCan":       res.get("cc_qualifies", "FAIL"),
                        "CC_Score":        res.get("cc_score", "—"),
                        "Rev_CAGR%":       res.get("cc_rev_cagr"),
                        "ROCE_Avg%":       res.get("cc_roce_avg"),
                        "Debt_Equity":     res.get("cc_debt_equity"),
                        "MCap_B_JPY":      res.get("market_cap_b_jpy"),
                        "Error":           res.get("error", ""),
                    }
                    fund_rows.append(fund_row)
                    if res.get("f_strong") and res.get("cc_qualifies") == "PASS":
                        triple_rows.append(fund_row.copy())
                    if done % 20 == 0 or done == len(breakout_tickers):
                        print(f"    {done}/{len(breakout_tickers)} done  "
                              f"(triple hits: {len(triple_rows)})")
                except Exception as e:
                    print(f"    {tkr}: error — {e}")
    else:
        print("\nStage 4 — Skipped")

    # Stage 5 — Excel
    print("\nStage 5 — Saving Excel workbook …")
    date_str = datetime.today().strftime("%Y%m%d_%H%M")
    path = DOWNLOAD_DIR / f"japan_market_scan_{date_str}.xlsx"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        def write_sheet(rows, name, sort_col=None):
            if not rows:
                pd.DataFrame().to_excel(writer, sheet_name=name, index=False)
                return
            df = pd.DataFrame(rows)
            if sort_col and sort_col in df.columns:
                df = df.sort_values(sort_col, ascending=False)
            df.to_excel(writer, sheet_name=name, index=False)
            if OPENPYXL_OK:
                style_sheet(writer.sheets[name])

        write_sheet(all_rows,    "All_Stocks",    sort_col="Change%")
        write_sheet(darvas_rows, "Darvas_Signals", sort_col="Upside_to_Top%")
        write_sheet(fund_rows,   "Fundamentals",  sort_col="Piotroski_Score")
        write_sheet(triple_rows, "Triple_Hits",   sort_col="Piotroski_Score")

    print(f"\n{'='*60}")
    print(f"  SCAN COMPLETE — {datetime.now().strftime('%d %b %Y  %H:%M:%S')}")
    print(f"  TSE stocks scanned:    {len(all_rows)}")
    print(f"  Darvas Breakouts:      {len(breakout_tickers)}")
    print(f"  Fundamentals scanned:  {len(fund_rows)}")
    print(f"  ★ TRIPLE HITS:         {len(triple_rows)}")
    if triple_rows:
        print("\n  Triple-hit stocks:")
        for r in sorted(triple_rows, key=lambda x: x.get("Piotroski_Score") or 0, reverse=True):
            print(f"    [{r['Code']}] {r['Name']:<30}  F={r['Piotroski_Score']}/9  "
                  f"CC={r['CC_Score']}  ¥{r['LTP_JPY']:,.0f}")
    else:
        print("\n  No Triple Hit stocks found today.")
    print(f"\n  📊 Excel saved → {path}\n")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
