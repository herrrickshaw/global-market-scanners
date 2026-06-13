# full_european_market_scan.py
# ===========================
# Optimized EURO STOXX 50 universe scanner (50 major Eurozone equity stocks).
# Features:
#   - Static Metadata: Instant Name & Sector mapping (zero API calls for index list).
#   - Restricts slow financials & info scans exclusively to breakout candidates.
#   - Auto-caches Ticker data to prevent duplicate HTTP calls.
#
# Pipeline:
#   Stage 1 & 2 — Bulk download 1-year OHLC for all 50 tickers in one batch.
#   Stage 3 — Compute Darvas Box & 200-day MA trend locally (instant).
#   Stage 4 — Run parallel fundamental scans (Piotroski + Coffee Can + PEGY) on breakouts only.
#   Stage 5 — Save a beautifully styled Excel workbook.
#
# Usage:
#   python full_european_market_scan.py
#   python full_european_market_scan.py --top 10      # Test run first 10
#   python full_european_market_scan.py --no-scans    # Skip fundamental scans
#
# Install:
#   pip install yfinance pandas openpyxl requests

import argparse
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import pandas as pd
import yfinance as yf

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Configuration & Constants ────────────────────────────────────────────────
DOWNLOAD_DIR = Path("./european_scan")
DOWNLOAD_DIR.mkdir(exist_ok=True)

DARVAS_CONFIRM   = 3
MAX_WORKERS       = 10
PIOTROSKI_STRONG  = 7

# Static Metadata for Euro Stoxx 50 components (removes 50 slow .info API calls)
EURO_STOXX_50_META = {
    "ADS.DE": ("Adidas", "Consumer Cyclical"),
    "ADYEN.AS": ("Adyen", "Financial Services"),
    "AD.AS": ("Ahold Delhaize", "Consumer Defensive"),
    "AI.PA": ("Air Liquide", "Basic Materials"),
    "AIR.PA": ("Airbus", "Industrials"),
    "ALV.DE": ("Allianz", "Financial Services"),
    "ABI.BR": ("Anheuser-Busch InBev", "Consumer Defensive"),
    "ARGX.BR": ("Argenx", "Healthcare"),
    "ASML.AS": ("ASML Holding", "Technology"),
    "CS.PA": ("Axa", "Financial Services"),
    "BAS.DE": ("BASF", "Basic Materials"),
    "BAYN.DE": ("Bayer", "Healthcare"),
    "BBVA.MC": ("BBVA", "Financial Services"),
    "SAN.MC": ("Banco Santander", "Financial Services"),
    "BMW.DE": ("BMW", "Consumer Cyclical"),
    "BNP.PA": ("BNP Paribas", "Financial Services"),
    "BN.PA": ("Danone", "Consumer Defensive"),
    "DBK.DE": ("Deutsche Bank", "Financial Services"),
    "DB1.DE": ("Deutsche Börse", "Financial Services"),
    "DHL.DE": ("Deutsche Post", "Industrials"),
    "DTE.DE": ("Deutsche Telekom", "Communication Services"),
    "ENEL.MI": ("Enel", "Utilities"),
    "ENI.MI": ("Eni", "Energy"),
    "EL.PA": ("EssilorLuxottica", "Healthcare"),
    "RACE.MI": ("Ferrari", "Consumer Cyclical"),
    "RMS.PA": ("Hermès", "Consumer Cyclical"),
    "IBE.MC": ("Iberdrola", "Utilities"),
    "ITX.MC": ("Inditex", "Consumer Cyclical"),
    "IFX.DE": ("Infineon Technologies", "Technology"),
    "INGA.AS": ("ING Group", "Financial Services"),
    "ISP.MI": ("Intesa Sanpaolo", "Financial Services"),
    "OR.PA": ("L'Oréal", "Consumer Defensive"),
    "MC.PA": ("LVMH", "Consumer Cyclical"),
    "MBG.DE": ("Mercedes-Benz Group", "Consumer Cyclical"),
    "MUV2.DE": ("Munich Re", "Financial Services"),
    "NDA-FI.HE": ("Nordea", "Financial Services"),
    "PRX.AS": ("Prosus", "Technology"),
    "RHM.DE": ("Rheinmetall", "Industrials"),
    "SAF.PA": ("Safran", "Industrials"),
    "SGO.PA": ("Saint-Gobain", "Industrials"),
    "SAN.PA": ("Sanofi", "Healthcare"),
    "SAP.DE": ("SAP", "Technology"),
    "SU.PA": ("Schneider Electric", "Industrials"),
    "SIE.DE": ("Siemens", "Industrials"),
    "ENR.DE": ("Siemens Energy", "Utilities"),
    "TTE.PA": ("TotalEnergies", "Energy"),
    "DG.PA": ("Vinci SA", "Industrials"),
    "UCG.MI": ("UniCredit", "Financial Services"),
    "VOW.DE": ("Volkswagen", "Consumer Cyclical"),
    "WKL.AS": ("Wolters Kluwer", "Professional Services")
}

# ── Financials Extractors ─────────────────────────────────────────────────────
def _first_df(ticker, *attrs):
    for attr in attrs:
        df = getattr(ticker, attr, None)
        if df is not None and isinstance(df, pd.DataFrame) and not df.empty:
            return df
    return None

def _row(df, *row_names, col: int = 0):
    if df is None or df.empty:
        return None
    for name in row_names:
        if name in df.index:
            try:
                val = df.loc[name].iloc[col]
                return float(val) if pd.notna(val) else None
            except Exception:
                pass
    return None

def _series(df, *rows):
    if df is None or df.empty:
        return []
    for name in rows:
        if name in df.index:
            return [float(v) for v in df.loc[name].dropna() if pd.notna(v)]
    return []

# ── Technical Scans ───────────────────────────────────────────────────────────
def compute_darvas_box(df: pd.DataFrame, confirm: int = DARVAS_CONFIRM) -> dict:
    """Detect Darvas Box from historical Close/High/Low series."""
    if df is None or df.empty or len(df) < confirm + 5:
        return {"signal": "INSUFFICIENT_DATA", "box_top": None, "box_bottom": None}

    highs  = pd.to_numeric(df["High"],  errors="coerce").fillna(0).tolist()
    lows   = pd.to_numeric(df["Low"],   errors="coerce").fillna(0).tolist()
    closes = pd.to_numeric(df["Close"], errors="coerce").fillna(0).tolist()

    current = closes[-1]
    highs_h = highs[:-1]  # Exclude current bar to avoid lookahead contamination
    lows_h  = lows[:-1]
    n       = len(highs_h)

    # Step 1: Box Top
    box_top_idx = box_top = None
    for i in range(n - confirm - 1, -1, -1):
        c = highs_h[i]
        if c == 0:
            continue
        w = highs_h[i + 1 : i + 1 + confirm]
        if len(w) == confirm and all(h < c for h in w):
            box_top_idx, box_top = i, c
            break

    if box_top is None:
        return {"signal": "NO_BOX", "box_top": None, "box_bottom": None}

    # Step 2: Box Bottom (historical segment starting from the box top)
    seg = lows_h[box_top_idx:]
    box_bottom = None
    for i in range(len(seg) - confirm):
        c = seg[i]
        if c == 0:
            continue
        w = seg[i + 1 : i + 1 + confirm]
        if len(w) == confirm and all(l > c for l in w):
            box_bottom = c
            break
    if box_bottom is None:
        valid = [l for l in seg if l > 0]
        box_bottom = min(valid) if valid else None

    if box_bottom is None:
        return {"signal": "NO_BOX", "box_top": round(box_top, 2), "box_bottom": None}

    # Step 3: Classify Signal
    if current > box_top:
        signal = "BREAKOUT_BUY"
    elif current < box_bottom:
        signal = "BREAKDOWN_SELL"
    else:
        signal = "IN_BOX"

    rng        = box_top - box_bottom
    upside_pct = (box_top - current) / current * 100 if current else 0
    pos_in_box = (current - box_bottom) / rng * 100 if rng else 0

    return {
        "signal":        signal,
        "box_top":       round(box_top,    2),
        "box_bottom":    round(box_bottom, 2),
        "current_price": round(current,    2),
        "upside_pct":    round(upside_pct, 2),
        "pos_in_box":    round(pos_in_box, 1),
    }

# ── Fundamental Scan ──────────────────────────────────────────────────────────
def fundamental_scan(symbol: str) -> dict:
    """Fetch financials and execute F-Score, Coffee Can, and PEGY calculations in one Ticker call."""
    res = {"symbol": symbol, "f_score": None, "cc_qualifies": "FAIL", "error": None}
    try:
        ticker = yf.Ticker(symbol)
        inc = _first_df(ticker, "income_stmt", "financials")
        bal = _first_df(ticker, "balance_sheet")
        cf  = _first_df(ticker, "cash_flow", "cashflow")
        
        info = {}
        try:
            info = ticker.info or {}
        except Exception:
            pass

        if inc is None or inc.empty:
            res["error"] = "no_financial_statements"
            return res
    except Exception as e:
        res["error"] = str(e)[:80]
        return res

    # ── 1. Piotroski F-Score ──────────────────────────────────────────────────
    try:
        ni0 = _row(inc, "Net Income", col=0);  a0 = _row(bal, "Total Assets", col=0)
        ni1 = _row(inc, "Net Income", col=1);  a1 = _row(bal, "Total Assets", col=1)
        roa0 = (ni0 / a0) if (ni0 and a0) else None
        roa1 = (ni1 / a1) if (ni1 and a1) else None
        ocf0 = _row(cf, "Operating Cash Flow", "Total Cash From Operating Activities", col=0)

        f1 = 1 if (roa0 and roa0 > 0) else 0
        f2 = 1 if (ocf0 and ocf0 > 0) else 0
        f3 = 1 if (roa0 and roa1 and roa0 > roa1) else 0
        f4 = 1 if (ocf0 and a0 and roa0 and (ocf0 / a0) > roa0) else 0

        ltd0 = _row(bal, "Long Term Debt", col=0) or 0
        ltd1 = _row(bal, "Long Term Debt", col=1) or 0
        f5 = 1 if (a0 and a1 and (ltd0/a0) < (ltd1/a1)) else 0

        ca0 = _row(bal, "Current Assets", "Total Current Assets", col=0)
        cl0 = _row(bal, "Current Liabilities", "Total Current Liabilities", col=0)
        ca1 = _row(bal, "Current Assets", "Total Current Assets", col=1)
        cl1 = _row(bal, "Current Liabilities", "Total Current Liabilities", col=1)
        f6 = 1 if (ca0 and cl0 and ca1 and cl1 and (ca0/cl0) > (ca1/cl1)) else 0

        sh0 = _row(bal, "Share Issued", col=0);  sh1 = _row(bal, "Share Issued", col=1)
        f7 = (1 if sh0 <= sh1 else 0) if (sh0 and sh1) else 1

        rev0 = _row(inc, "Total Revenue", col=0);  gp0 = _row(inc, "Gross Profit", col=0)
        rev1 = _row(inc, "Total Revenue", col=1);  gp1 = _row(inc, "Gross Profit", col=1)
        f8 = 1 if (gp0 and rev0 and gp1 and rev1 and (gp0/rev0) > (gp1/rev1)) else 0
        f9 = 1 if (rev0 and a0 and rev1 and a1 and (rev0/a0) > (rev1/a1)) else 0

        res["f_score"] = f1 + f2 + f3 + f4 + f5 + f6 + f7 + f8 + f9
    except Exception:
        pass

    # ── 2. Coffee Can Screen ──────────────────────────────────────────────────
    try:
        c = {}
        revs = _series(inc, "Total Revenue")
        cagr = ((revs[0] / revs[-1]) ** (1 / (len(revs) - 1)) - 1) * 100 if len(revs) >= 2 and revs[-1] > 0 else None
        c["C1"] = 1 if (cagr and cagr > 10) else 0

        ebit_s = _series(inc, "EBIT", "Operating Income")
        ta_s   = _series(bal, "Total Assets")
        cl_s   = _series(bal, "Current Liabilities", "Total Current Liabilities")
        roce_l = [ebit_s[i] / (ta_s[i] - cl_s[i]) * 100 for i in range(min(len(ebit_s), len(ta_s), len(cl_s))) if (ta_s[i] - cl_s[i]) > 0]
        avg_roce = sum(roce_l) / len(roce_l) if roce_l else None
        c["C2"] = 1 if (avg_roce and avg_roce > 15) else 0

        de_raw = info.get("debtToEquity")
        de = de_raw / 100 if de_raw is not None and de_raw > 10 else de_raw
        if de is None:
            ltd_s = _series(bal, "Long Term Debt")
            eq_s  = _series(bal, "Stockholders Equity", "Total Stockholder Equity")
            de = (ltd_s[0] / abs(eq_s[0])) if (ltd_s and eq_s and eq_s[0] != 0) else None
        c["C3"] = 1 if (de is not None and de < 1) else 0

        mcap = info.get("marketCap") or getattr(ticker.fast_info, "market_cap", 0)
        c["C4"] = 1 if mcap >= 1e9 else 0  # >= 1 Billion EUR

        ni_s = _series(inc, "Net Income")
        c["C5"] = 1 if (ni_s and all(n > 0 for n in ni_s)) else 0

        fcf_s = _series(cf, "Free Cash Flow")
        if fcf_s:
            c["C6"] = 1 if fcf_s[0] > 0 else 0
        else:
            ocf_s, capex_s = _series(cf, "Operating Cash Flow"), _series(cf, "Capital Expenditure")
            c["C6"] = 1 if (ocf_s and capex_s and (ocf_s[0] - abs(capex_s[0])) > 0) else 0

        qualifies = sum(c.values()) == len(c)
        res.update({
            "cc_qualifies":   "PASS" if qualifies else "FAIL",
            "cc_score":       f"{sum(c.values())}/{len(c)}",
            "cc_rev_cagr":    round(cagr, 2) if cagr else None,
            "cc_roce_avg":    round(avg_roce, 2) if avg_roce else None,
            "cc_debt_equity": round(de, 2) if de else None,
        })
    except Exception:
        pass

    # ── 3. Valuation & PEGY ───────────────────────────────────────────────────
    try:
        pe  = info.get("trailingPE") or info.get("forwardPE") or info.get("regularMarketPrice", 0) / (info.get("trailingEps") or 1)
        peg = info.get("pegRatio")
        dy_raw = info.get("dividendYield")
        dy = (dy_raw if dy_raw >= 0.15 else dy_raw * 100) if dy_raw is not None else 0.0
        pegy = float(peg) + dy if peg is not None else None

        res.update({
            "market_cap_b":    round(mcap / 1e9, 2) if mcap else None,
            "pe_ratio":        round(pe, 2) if pe and pe > 0 else None,
            "pb_ratio":        round(info.get("priceToBook"), 2) if info.get("priceToBook") else None,
            "dividend_yield":  round(dy, 2),
            "pegy_ratio":      round(pegy, 2) if pegy else None,
        })
    except Exception:
        pass

    return res

# ── Sheet Styling ─────────────────────────────────────────────────────────────
def style_excel_sheet(ws):
    """Format and style an openpyxl sheet to premium standards."""
    if not OPENPYXL_OK:
        return
    
    fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    fill_zebra  = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body   = Font(name="Calibri", size=11, bold=False, color="000000")
    
    border_thin = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Style Header
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # Style Data Rows
    for row_idx in range(2, ws.max_row + 1):
        is_zebra = (row_idx % 2 == 1)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_body
            cell.border = border_thin
            if is_zebra:
                cell.fill = fill_zebra

            header = str(ws.cell(row=1, column=col_idx).value or "").upper()
            val = cell.value

            # Alignment
            if any(k in header for k in ["SYMBOL", "SIGNAL", "SCORE", "CLASS", "QUALIFIES", "EXCHANGE"]):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Formats
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if "%" in header or "YIELD" in header or "CAGR" in header or "ROCE" in header:
                    cell.number_format = '0.00"%"'
                elif "CAP" in header or "LTP" in header or "BOX" in header or "PRICE" in header or "200MA" in header:
                    cell.number_format = '#,##0.00'
                elif "PEG" in header or "PE_RATIO" in header or "PB_RATIO" in header or "DEBT_EQUITY" in header:
                    cell.number_format = '0.00'

    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20

    # Auto Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws.views.sheetView[0].showGridLines = True

# ── Main Scanner ──────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Optimized Euro Stoxx 50 Scanner")
    parser.add_argument("--top", type=int, default=0, help="Scan only first N tickers")
    parser.add_argument("--no-scans", action="store_true", default=False, help="Skip Stage 4 fundamental scans")
    args = parser.parse_args()

    symbols = list(EURO_STOXX_50_META.keys())
    if args.top:
        symbols = symbols[:args.top]

    print(f"\n{'#'*60}")
    print(f"  OPTIMIZED EURO STOXX 50 SCANNER")
    print(f"  Started: {datetime.now().strftime('%d %b %Y  %H:%M:%S')}")
    print(f"{'#'*60}")

    # Stage 1 & 2: Bulk OHLC download (1 request for all 50 tickers - extremely fast)
    print(f"\nStage 1 & 2 — Bulk downloading 1y history for {len(symbols)} stocks …")
    try:
        raw_ohlc = yf.download(symbols, period="1y", group_by="ticker", auto_adjust=True, threads=True, progress=False)
    except Exception as e:
        sys.exit(f"❌ Failed to download bulk OHLC data: {e}")

    # Stage 3: Technical Scan (Local pandas math - instant)
    print("\nStage 3 — Computing Darvas Box & 200-day MA Trend locally …")
    all_stocks_map = {}
    breakout_symbols = []

    for sym in symbols:
        try:
            df = raw_ohlc[sym].dropna() if isinstance(raw_ohlc.columns, pd.MultiIndex) else raw_ohlc.dropna()
            if df.empty or len(df) < 20:
                continue

            closes  = pd.to_numeric(df["Close"], errors="coerce").dropna()
            ltp     = round(float(closes.iloc[-1]), 2)
            prev    = round(float(closes.iloc[-2]), 2) if len(closes) >= 2 else None
            chg_pct = round((ltp - prev) / prev * 100, 2) if prev else None

            # 200-day Moving Average
            ma_200 = round(closes.rolling(200).mean().iloc[-1], 2) if len(closes) >= 200 else None
            dist_ma = round(((ltp - ma_200) / ma_200 * 100), 2) if ma_200 else None
            trend_sig = "Above 200MA (Uptrend)" if dist_ma and dist_ma > 5 else "Below 200MA (Downtrend)" if dist_ma and dist_ma < -5 else "Near 200MA (Consolidation)" if dist_ma else "Insufficient History"

            # Darvas Box
            darv = compute_darvas_box(df)
            name, sector = EURO_STOXX_50_META.get(sym, (sym, "—"))

            all_stocks_map[sym] = {
                "Symbol":             sym,
                "Name":               name,
                "Sector":             sector,
                "LTP":                ltp,
                "Change%":            chg_pct,
                "200_Day_MA":         ma_200,
                "Distance_to_200MA%": dist_ma,
                "Trend_Signal":       trend_sig,
                "Darvas_Signal":      darv.get("signal"),
                "Box_Top":            darv.get("box_top"),
                "Box_Bottom":         darv.get("box_bottom"),
                "Upside_to_Top%":     darv.get("upside_pct"),
                "Position_in_Box%":   darv.get("pos_in_box"),
                # Valuation fallbacks
                "Market_Cap_B":       None,
                "P_E_Ratio":          None,
                "P_B_Ratio":          None,
                "Dividend_Yield%":    0.0,
            }

            if darv.get("signal") == "BREAKOUT_BUY":
                breakout_symbols.append(sym)
        except Exception as e:
            print(f"  ⚠️ Technical error on {sym}: {e}")

    print(f"  Processed {len(all_stocks_map)} stocks. Found {len(breakout_symbols)} Darvas breakouts.")

    # Stage 4: Fundamental scans (Only run on breakout candidates - huge speedup)
    fund_rows = []
    triple_hits = []

    if not args.no_scans and breakout_symbols:
        print(f"\nStage 4 — Fetching financials & scans for {len(breakout_symbols)} breakouts ({MAX_WORKERS} workers) …")
        done = 0
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
            futures = {pool.submit(fundamental_scan, sym): sym for sym in breakout_symbols}
            for future in as_completed(futures):
                sym = futures[future]
                done += 1
                try:
                    res = future.result()
                    tech = all_stocks_map.get(sym, {})
                    
                    # Merge technicals with fundamental scan result
                    fund_row = {
                        "Symbol":             sym,
                        "Name":               tech.get("Name"),
                        "Sector":             tech.get("Sector"),
                        "LTP":                tech.get("LTP"),
                        "Change%":            tech.get("Change%"),
                        "Darvas_Signal":      tech.get("Darvas_Signal"),
                        "200_Day_MA":         tech.get("200_Day_MA"),
                        "Distance_to_200MA%": tech.get("Distance_to_200MA%"),
                        "P_E_Ratio":          res.get("pe_ratio"),
                        "P_B_Ratio":          res.get("pb_ratio"),
                        "Dividend_Yield%":    res.get("dividend_yield"),
                        "PEGY_Ratio":         res.get("pegy_ratio"),
                        "Piotroski_Score":    res.get("f_score"),
                        "Piotroski_Class":    res.get("f_interpretation"),
                        "CoffeeCan_Score":    res.get("cc_score"),
                        "CoffeeCan_Class":    res.get("cc_qualifies"),
                        "CAGR_Revenue%":      res.get("cc_rev_cagr"),
                        "ROCE_Average%":      res.get("cc_roce_avg"),
                        "Debt_Equity":        res.get("cc_debt_equity"),
                        "Error":              res.get("error")
                    }
                    fund_rows.append(fund_row)

                    # Update the master mapping so All_Stocks sheet contains PE/PB/Yield for breakouts
                    tech.update({
                        "Market_Cap_B":    res.get("market_cap_b"),
                        "P_E_Ratio":       res.get("pe_ratio"),
                        "P_B_Ratio":       res.get("pb_ratio"),
                        "Dividend_Yield%": res.get("dividend_yield"),
                    })

                    if res.get("f_score") and res.get("f_score") >= PIOTROSKI_STRONG and res.get("cc_qualifies") == "PASS":
                        triple_hits.append(fund_row.copy())

                    if done % 5 == 0 or done == len(breakout_symbols):
                        print(f"    {done}/{len(breakout_symbols)} processed (Triple Hits so far: {len(triple_hits)})")
                except Exception as e:
                    print(f"    ❌ Fundamental error on {sym}: {e}")
    else:
        print("\nStage 4 — Skipped (No active breakouts or --no-scans set)")

    # Stage 5: Save formatted Excel
    print("\nStage 5 — Saving Excel workbook …")
    all_stocks = list(all_stocks_map.values())
    darvas_signals = [s for s in all_stocks if s["Darvas_Signal"] in ("BREAKOUT_BUY", "BREAKDOWN_SELL")]

    date_str = datetime.today().strftime("%Y%m%d_%H%M")
    excel_path = DOWNLOAD_DIR / f"european_market_scan_{date_str}.xlsx"

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        def write_sheet(rows, name, sort_col=None):
            if not rows:
                pd.DataFrame().to_excel(writer, sheet_name=name, index=False)
                return
            df = pd.DataFrame(rows)
            if sort_col and sort_col in df.columns:
                df = df.sort_values(sort_col, ascending=False)
            df.to_excel(writer, sheet_name=name, index=False)
            if OPENPYXL_OK:
                style_excel_sheet(writer.sheets[name])

        write_sheet(all_stocks,    "All_Stocks",     sort_col="Change%")
        write_sheet(darvas_signals, "Darvas_Signals",  sort_col="Upside_to_Top%")
        write_sheet(fund_rows,   "Fundamentals",    sort_col="Piotroski_Score")
        write_sheet(triple_hits, "Triple_Hits",     sort_col="Piotroski_Score")

    print(f"\n{'='*60}")
    print(f"  SCAN COMPLETE — {datetime.now().strftime('%d %b %Y  %H:%M:%S')}")
    print(f"  Total European Stocks Scanned: {len(all_stocks)}")
    print(f"  Darvas Breakouts Found:       {len(breakout_symbols)}")
    print(f"  Fundamentals Scanned:         {len(fund_rows)}")
    print(f"  ★ TRIPLE HITS FOUND:          {len(triple_hits)}")
    print(f"{'='*60}")

    if triple_hits:
        print("\n  Triple Hit Stocks:")
        for r in sorted(triple_hits, key=lambda x: x.get("Piotroski_Score") or 0, reverse=True):
            print(f"    • {r['Symbol']:<10} {r['Name']:<30} F-Score: {r['Piotroski_Score']}/9  CC: {r['CoffeeCan_Score']}  LTP: €{r['LTP']}")
    else:
        print("\n  No Triple Hit stocks found today.")

    print(f"\n  📊 Excel report saved → {excel_path}\n")

if __name__ == "__main__":
    main()
